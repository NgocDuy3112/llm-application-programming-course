from pathlib import Path

from datasets import Dataset
from openpyxl import load_workbook

from src.utils.text import extract_answer


def _resolve_dataset_path(dataset_path: str) -> Path:
    path = Path(dataset_path)
    if path.is_absolute():
        return path
    return Path(__file__).resolve().parents[2] / path


def load_dataset_from_excel(dataset_path: str):
    """Tải bộ dữ liệu Excel cục bộ và chuẩn hóa các cột của nó."""
    path = _resolve_dataset_path(dataset_path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if "query_vi" not in headers:
        raise ValueError(f"Missing required column 'query_vi' in {path}")

    response_header = next(
        (h for h in headers if isinstance(h, str) and h.startswith("response_vi")),
        None,
    )
    if response_header is None:
        raise ValueError(f"Missing required response column in {path}")

    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {}
        for header, value in zip(headers, values):
            if header == "query_vi":
                row["query_vi"] = value
            elif header == response_header:
                row["response_vi"] = value
        if row:
            rows.append(row)

    wb.close()
    return Dataset.from_list(rows)


def add_ground_truth(ds):
    return ds.map(
        lambda example: {"ground_truth": extract_answer(example["response_vi"])},
        remove_columns=[],
    )
