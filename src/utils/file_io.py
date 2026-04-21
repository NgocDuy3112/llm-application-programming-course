"""File I/O utilities: save results to Excel and CSV."""
import csv

import openpyxl


def save_dataset_to_excel(ds, path: str):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="dataset")
    headers = list(ds.column_names)
    ws.append(headers)
    for row in ds:
        ws.append([row.get(col, "") for col in headers])
    wb.save(path)
    print(f"Saved dataset to {path}")


def export_query_response(input_path: str, output_path: str):
    """Read an Excel file and export only query_vi and response_vi columns."""
    wb_in = openpyxl.load_workbook(input_path, read_only=True)
    ws_in = wb_in.active
    headers = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    query_col = next((i for i, h in enumerate(headers) if h == "query_vi"), None)
    response_col = next((i for i, h in enumerate(headers) if isinstance(h, str) and h.startswith("response_vi")), None)
    if query_col is None or response_col is None:
        raise ValueError(f"Columns not found. Available: {headers}")
    wb_out = openpyxl.Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title="data")
    ws_out.append(["query_vi", "response_vi"])
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        ws_out.append([row[query_col], row[response_col]])
    wb_in.close()
    wb_out.save(output_path)
    print(f"Exported to {output_path}")


def save_accuracy_to_csv(records, path: str):
    with open(path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "batch_idx",
                "batch_start",
                "batch_end",
                "model",
                "total",
                "correct_no_cot",
                "correct_cot",
                "accuracy_no_cot",
                "accuracy_cot",
                "accuracy_diff",
            ],
        )
        writer.writeheader()
        writer.writerows(records)
    print(f"Saved accuracy CSV to {path}")
